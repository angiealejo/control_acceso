# coding=utf-8
import hashlib

from django.contrib.auth.models import User
from django.db import transaction
from django.http import QueryDict, HttpResponse

from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, PatternFill

from apps.comun.forms import DireccionForm
from apps.comun.models import pais_default
from apps.comun.utilidades import convertirnumeroempleadoentero, traergrupos, traerpuntoscontrol
from apps.empleado.forms import RegistrarEditarVariosForm, EmpleadoForm
from apps.empleado.models import DatosUsuarioEmpleado
from apps.usuario.models import PasswordCliente


class GuardarEmpleadoMixin:

        def __init__(self):
            pass

        @transaction.atomic
        def guardar(self, fila, request):
            tsp = transaction.savepoint()
            pais = pais_default
            error = {u'empleado_errores': None, u'datosusuarioempleado_errores': None, u'errores': None,
                     u'NoGuardado': False}
            errores = u''
            errores_empleado = u''
            errores_direccion = u''
            try:
                numero_empleado = None
                if fila[0] is not None:
                    numero_empleado = convertirnumeroempleadoentero(int(fila[0]))
                sinonulos = [1, 2, 5, 6, 7]
                for f in sinonulos:  # Campos requeridos que si traen algo se vuelvan unicode
                    if fila[f] is not None:
                        fila[f] = unicode(fila[f])
                sinulos = [3, 4, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]
                for f in sinulos:  # Campos que pueden ser nulos pero es mejor que esten blancos
                    if fila[f] is None:
                        fila[f] = u''
                    else:
                        fila[f] = unicode(fila[f])
                if fila[7] is not None:  # Fecha de nacimiento cortando para que solo quede yyyy-mm-dd
                    fila[7] = fila[7][:10]
                try:
                    instancia = DatosUsuarioEmpleado.objects.get(numero_empleado__iexact=numero_empleado)
                    instancia_empleado = instancia.empleado
                    instancia_direccion = instancia.empleado.direccion
                    datos = RegistrarEditarVariosForm(instance=instancia)
                    empleado = EmpleadoForm(instance=instancia_empleado)
                    direccion = DireccionForm(instance=instancia_direccion)
                    usuario = instancia.usuario
                except Exception:
                    datos = RegistrarEditarVariosForm()
                    empleado = EmpleadoForm()
                    direccion = DireccionForm()
                    usuario = None
                formsdict = {
                    u'csrfmiddlewaretoken': request.POST[u'csrfmiddlewaretoken'],
                    u'numero_empleado_entero': fila[0],
                    u'numero_empleado': numero_empleado,
                    u'email': fila[4],
                    u'hora_entrada': fila[5],
                    u'hora_salida': fila[6],
                    u'nombre': fila[1],
                    u'apellido_paterno': fila[2],
                    u'apellido_materno': fila[3],
                    u'fecha_nacimiento': fila[7],
                    u'curp': fila[8],
                    u'rfc': fila[9],
                    u'pais': fila[10],
                    u'estado': fila[11],
                    u'municipio': fila[12],
                    u'ciudad': fila[13],
                    u'asentamiento': fila[14],
                    u'calle': fila[15],
                    u'numero_exterior': fila[16],
                    u'numero_interior': fila[17],
                    u'codigo_postal': fila[18],
                    u'datos_adicionales': fila[19]
                }
                formsqdict = QueryDict('', mutable=True, encoding='utf-8')
                formsqdict.update(formsdict)
                empleado_verificar_remplazo = EmpleadoForm()
                direccion_verificar_remplazo = DireccionForm()
                empleado_verificar_remplazo.cleaned_data = empleado_verificar_remplazo.data = formsqdict
                direccion_verificar_remplazo.cleaned_data = direccion_verificar_remplazo.data = formsqdict
                empleado_verificar_remplazo.is_bound = direccion_verificar_remplazo.is_bound = True
                reemplazar = False
                if not empleado_verificar_remplazo.is_valid():
                    errores_empleado = unicode(empleado_verificar_remplazo.errors)
                    formsdict = self.reemplazarempleado(empleado_verificar_remplazo, formsdict)
                    reemplazar = True
                if not direccion_verificar_remplazo.is_valid():
                    errores_direccion = unicode(direccion_verificar_remplazo.errors)
                    formsdict = self.reemplazardireccion(direccion_verificar_remplazo, formsdict)
                    reemplazar = True
                if reemplazar:
                    errores += errores_empleado + errores_direccion
                    formsqdict = QueryDict('', mutable=True, encoding='utf-8')
                    formsqdict.update(formsdict)
                datos.cleaned_data = datos.data = formsqdict
                empleado.cleaned_data = empleado.data = formsqdict
                direccion.cleaned_data = direccion.data = formsqdict
                datos.is_bound = True
                empleado.is_bound = True
                direccion.is_bound = True
                if empleado.is_valid() and direccion.is_valid() and datos.is_valid():
                    empleado = empleado.save()
                    direccion = direccion.save()
                    direccion.pais = pais
                    direccion.save()
                    empleado.direccion = direccion
                    empleado.save()
                    datos = datos.save()
                    if usuario is None:
                        usuario = User(username=str(int(numero_empleado)), email=fila[4])
                        usuario.set_password(numero_empleado)
                        usuario.save()
                        passwordhasheada = hashlib.md5(numero_empleado).hexdigest()
                        password = PasswordCliente(password=passwordhasheada)
                        password.save()
                        datos.password = password
                        grupos = traergrupos(0)
                        usuario.groups.clear()
                        for grupo in grupos:
                            usuario.groups.add(grupo)
                        usuario.save()
                    else:
                        usuario.email = fila[4]
                        usuario.save()
                    usuario.first_name = empleado.nombre
                    if empleado.apellido_materno == '':
                        usuario.last_name = empleado.apellido_paterno
                    else:
                        usuario.last_name = empleado.apellido_paterno + ' ' + empleado.apellido_materno
                    usuario.save()
                    datos.usuario = usuario
                    datos.empleado = empleado
                    datos.save()
                    puntoscontrol = traerpuntoscontrol(request.POST.getlist(u'puntocontrol'))
                    for puntocontrol in puntoscontrol:
                        datos.puntocontrol.add(puntocontrol)
                    datos.save()
                    if tsp:
                        transaction.savepoint_commit(tsp)
                else:
                    error[u'NoGuardado'] = True
                    if not empleado.is_valid():
                        error[u'empleado_errores'] = empleado
                        error_empleado = unicode(empleado.errors)
                        if errores_empleado == u'':
                            errores += error_empleado
                    if not direccion.is_valid():
                        error_direccion = unicode(direccion.errors)
                        if errores_direccion == u'':
                            errores += error_direccion
                    if not datos.is_valid():
                        error[u'datosusuarioempleado_errores'] = datos
                        errores += unicode(datos.errors)
            except Exception, e:
                errores += u'<ul class="errorlist">' \
                           u'<li>Error interno del servidor' \
                           u'<ul class="errorlist">' \
                           u'<li>' + unicode(e.message) +\
                           u'</li></ul></li></ul>'
            error[u'errores'] = self.nombreserrores(errores)
            return error

        @staticmethod
        def resultado(self, contador, guardado, fila, error):
            clase = u'class="danger"'
            color = u'style="color: red;"'
            if guardado == u'creado':
                clase = u'class="success"'
                color = u'style="color: green;"'
            elif guardado == u'modificado':
                clase = u'class="info"'
                color = u'style="color: blue;"'
            idresultado = u'id="' + unicode(contador)+u'_resultado"'
            resultado = u'<tr ' + idresultado + u' ' + clase + u'>'
            celdas_resultado = [u'', u'', u'', u'', u'', u'', u'']
            idfilas = [0, 1, 2, 3, 4, 5, 6]
            for idfila in idfilas:
                if fila[idfila] is not None and fila[idfila] != u'' and fila[idfila] != '':
                    celdas_resultado[idfila] = unicode(fila[idfila])
            celdas_resultado_revisadas = self.chekarerrores(celdas_resultado, error)
            for celda in celdas_resultado_revisadas:
                resultado += u'<td><strong><span ' + color + u'>' + celda + u'</span></strong></td>'
            resultado += u'</tr>'
            return resultado

        @staticmethod
        def reemplazarempleado(empleado, fila):
            campos = [u'fecha_nacimiento', u'curp', u'rfc']
            for campo in campos:
                reemplazo = None
                if campo != u'fecha_nacimiento':
                    reemplazo = u''
                if empleado.has_error(campo):
                    fila[campo] = reemplazo
            return fila

        @staticmethod
        def reemplazardireccion(direccion, fila):
            campos = [u'pais', u'estado', u'municipio', u'ciudad', u'calle', u'asentamiento',
                      u'numero_interior', u'numero_exterior', u'codigo_postal', u'datos_adicionales']
            for campo in campos:
                if direccion.has_error(campo):
                    fila[campo] = u''
            return fila

        @staticmethod
        def chekarerrores(celdas_resultado, error):
            empleado = error[u'empleado_errores']
            datos = error[u'datosusuarioempleado_errores']
            if empleado is not None:
                campos = [u'nombre', u'apellido_paterno', u'apellido_materno']
                celdas = [1, 2, 3]
                contador = 0
                while contador < len(celdas):
                    if empleado.has_error(campos[contador]):
                        celdas_resultado[celdas[contador]] = u''
                        errores = 0
                        while errores < len(empleado.errors[campos[contador]]):
                            celdas_resultado[celdas[contador]] += empleado.errors[campos[contador]][errores]
                            errores += 1
                    contador += 1
            if datos is not None:
                campos = [u'numero_empleado', u'email', u'hora_entrada', u'hora_salida']
                celdas = [0, 4, 5, 6]
                contador = 0
                while contador < len(celdas):
                    if datos.has_error(campos[contador]):
                        celdas_resultado[celdas[contador]] = u''
                        errores = 0
                        while errores < len(datos.errors[campos[contador]]):
                            celdas_resultado[celdas[contador]] += datos.errors[campos[contador]][errores]
                            errores += 1
                    contador += 1
            return celdas_resultado

        @staticmethod
        def nombreserrores(errores):
            campos = [u'puntocontrol', u'hora_entrada', u'hora_salida', u'numero_empleado', u'email',
                      u'pais', u'estado', u'municipio', u'ciudad', u'calle', u'asentamiento',
                      u'numero_interior', u'numero_exterior', u'codigo_postal', u'datos_adicionales',
                      u'nombre', u'apellido_paterno', u'apellido_materno', u'fecha_nacimiento', u'curp', u'rfc']
            campos_correctos = [u'Puntos de Control:', u'Hora de Entrada:', u'Hora de Salida:', u'Número de Empleado:',
                                u'Correo Electrónico:', u'País:', u'Estado:', u'Municipio:', u'Ciudad:', u'Calle:',
                                u'Asentamiento:', u'Número Interno:', u'Número Externo:', u'Codigo Postal:',
                                u'Datos Adicionales:', u'Nombre:', u'Apellido Paterno:', u'Apellido Materno:',
                                u'Fecha de Nacimiento:', u'CURP:', u'RFC:']
            contador = 0
            while contador < len(campos):
                original = u'<li>' + campos[contador] + u'<ul'
                corregido = u'<li>' + campos_correctos[contador] + u'<ul'
                errores = errores.replace(original, corregido)
                contador += 1
            return unicode(errores)


class GenerarExcelMixin:

        def __init__(self):
            pass

        @staticmethod
        def generarexcelresponse(nombre_archivo_excel, columnas_ancho, fila_titulo, consulta):
            workbook = Workbook()
            worksheet = workbook.get_active_sheet()
            worksheet.title = nombre_archivo_excel
            row_num = 0
            for col_num in xrange(len(columnas_ancho)):
                c = worksheet.cell(row=row_num + 1, column=col_num + 1)
                c.value = columnas_ancho[col_num][0]
                c.style.font.bold = True
                worksheet.column_dimensions[get_column_letter(col_num + 1)].width = columnas_ancho[col_num][1]
            relleno = PatternFill(fill_type='solid', start_color='A9A9F5')
            fuente = Font(name='Arial Bold', size=12)
            for row in worksheet.iter_rows(fila_titulo):
                for cell in row:
                    cell.font = fuente
                    cell.fill = relleno
            for elemento in consulta:
                row_num += 1
                row = []
                for elemento_consulta in consulta.field_names:
                    row.append(elemento[elemento_consulta])
                for col_num in xrange(len(row)):
                    c = worksheet.cell(row=row_num + 1, column=col_num + 1)
                    c.value = row[col_num]
                    c.style.alignment.wrap_text = True
            response = HttpResponse(content=save_virtual_workbook(workbook),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=' + nombre_archivo_excel + '.xlsx'
            return response
