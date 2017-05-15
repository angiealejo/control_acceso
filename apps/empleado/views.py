# coding=utf-8
import hashlib
import json

from django.core.exceptions import SuspiciousOperation
from django.db import transaction
from django.contrib.auth.models import User
from django.core.urlresolvers import reverse_lazy
from django.db.models import Q
from django.http import HttpResponseRedirect, HttpResponse, Http404
from django.views.generic import TemplateView, FormView, View

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from openpyxl import load_workbook

from apps.bitacora.models import Bitacora
from apps.comun.forms import DireccionForm
from apps.comun.models import Direccion, pais_default
from apps.configuracion.models import Configuracion
from apps.empleado.forms import EmpleadoForm, RegistrarEditarForm, ImportarExcelForm
from apps.empleado.mixins import GuardarEmpleadoMixin, GenerarExcelMixin
from apps.empleado.models import Empleado, DatosUsuarioEmpleado
from apps.empleado.serializers import DatosSerializer, DatosClienteSerializer
from apps.comun.utilidades import traergrupos, convertirnumeroempleadoentero
from apps.punto_control.models import PuntoControl
from apps.usuario.models import PasswordCliente


class DeleteAPI(APIView):
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    activo = [0, False]

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        id_empleado = request.POST['id']
        response_data = {'exito': True}
        tsp = transaction.savepoint()
        try:
            datos = DatosUsuarioEmpleado.objects.get(id=id_empleado)
            if self.request.user != datos.usuario:
                if datos.activo == self.activo[0]:
                    response_data['exito'] = False
                else:
                    usuario = User.objects.get(id=datos.usuario.id)
                    usuario.is_active = self.activo[1]
                    usuario.save()
                    datos.activo = self.activo[0]
                    datos.save()
                if tsp:
                    transaction.savepoint_commit(tsp)
            else:
                response_data['exito'] = False
        except DatosUsuarioEmpleado.DoesNotExist:
            response_data['exito'] = False
        except Empleado.DoesNotExist:
            response_data['exito'] = False
        except User.DoesNotExist:
            response_data['exito'] = False
        return HttpResponse(json.dumps(response_data), content_type="application/json")


class RestoreAPI(DeleteAPI):
    activo = [1, True]


class DatosAPI(APIView):
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    def get(self, request, format=None):
        num_empleado_usuario = self.request.user.username
        empleado = request.GET.get('empleado')
        lista_datos = request.GET.get('lista')
        activo = request.GET.get('activo')
        usuario = request.GET.get('usuario')
        busqueda = request.GET.get('term[term]')
        configuracion = bool(request.GET.get('configuracion'))
        # Cuando se usa desde el cliente
        idpcrecibida = request.GET.get('idpuntocontrol')
        if busqueda is not None:
            datos = DatosUsuarioEmpleado.objects.filter(
                Q(empleado__nombre__icontains=busqueda) |
                Q(empleado__apellido_paterno__icontains=busqueda) |
                Q(empleado__apellido_materno__icontains=busqueda) |
                Q(numero_empleado__icontains=busqueda)
            ).order_by('numero_empleado')
            if configuracion:
                configuraciones = Configuracion.objects.all().filter(servidor=0)
                for empleado_configuracion in configuraciones:
                    datos = datos.exclude(id=empleado_configuracion.empleado.id)
            serializer = DatosClienteSerializer(datos, many=True)
        elif idpcrecibida is not None:
            puntocontrol = PuntoControl.objects.get(id=idpcrecibida, asignado=1, activo=1)
            datos = DatosUsuarioEmpleado.objects.filter(activo=1, puntocontrol=puntocontrol).order_by('numero_empleado')
            serializer = DatosClienteSerializer(datos, many=True)
        elif lista_datos is not None and lista_datos == "Todos" and activo is not None:
            datos = DatosUsuarioEmpleado.objects.filter(activo=activo).order_by('numero_empleado')
            serializer = DatosSerializer(datos, many=True)
        elif lista_datos is not None and activo is not None:
            datos = DatosUsuarioEmpleado.objects.filter(activo=activo).order_by('numero_empleado'). \
                exclude(usuario__username__exact=num_empleado_usuario)
            serializer = DatosSerializer(datos, many=True)
        elif empleado is not None:
            empleados = DatosUsuarioEmpleado.objects.get(id=empleado)
            serializer = DatosSerializer(empleados)
        elif usuario is not None:
            empleados = DatosUsuarioEmpleado.objects.get(usuario__username__exact=num_empleado_usuario)
            serializer = DatosSerializer(empleados)
        return HttpResponse(json.dumps(serializer.data), content_type="application/json")


class DatosActivos(View):
    # En otras consultas diferentes se deben sustituir estos datos y el metodo getfiltro
    activo = 1
    modelo = DatosUsuarioEmpleado
    columnas = ['usuario__username', 'empleado__nombre', 'empleado__apellido_paterno', 'empleado__apellido_materno',
                'empleado__fecha_nacimiento', 'empleado__curp', 'empleado__rfc', 'empleado__curp',
                'numero_empleado', 'numero_empleado']
    serializer = DatosSerializer

    def getfiltro(self, consulta, busqueda):
        return consulta.filter(
            Q(empleado__nombre__icontains=busqueda) |
            Q(empleado__apellido_paterno__icontains=busqueda) |
            Q(empleado__apellido_materno__icontains=busqueda) |
            Q(empleado__curp__icontains=busqueda) |
            Q(empleado__rfc__icontains=busqueda) |
            Q(empleado__fecha_nacimiento__icontains=busqueda) |
            Q(numero_empleado__icontains=busqueda)
        )

    def getconsulta(self):
        if self.modelo is DatosUsuarioEmpleado:
            consulta = self.modelo.objects.all().filter(activo=self.activo)
        elif self.modelo is Bitacora:
            if self.activo is 1:
                consulta = self.modelo.objects.all() \
                    .filter(empleado__usuario__username__exact=self.request.user.username)
            else:
                consulta = self.modelo.objects.all()
        elif self.modelo is Configuracion:
            consulta = self.modelo.objects.all().filter(servidor=self.activo)
        elif self.activo is not None:
            consulta = self.modelo.objects.all().filter(activo=self.activo)
        else:
            consulta = self.modelo.objects.all()
        return consulta

    def get(self, request, *args, **kwargs):
        consulta = self.getconsulta()
        iTotalRecords = consulta.count()
        columna = int(request.GET.get('iSortCol_0'))
        if request.GET.get('sSortDir_0') == 'desc':
            orden = '-' + self.columnas[columna]
        else:
            orden = self.columnas[columna]
        consulta = consulta.order_by(orden)
        sEcho = 0
        if request.GET.get('sEcho'):
            sEcho = int(request.GET.get('sEcho'))
        iTotalDisplayRecords = iTotalRecords
        if request.GET.get('sSearch') is not None and request.GET.get('sSearch') is not '':
            search = request.GET.get('sSearch')
            consulta = self.getfiltro(consulta, search)
            iTotalDisplayRecords = consulta.count()
        if request.GET.get('iDisplayStart'):
            iDisplayStart = request.GET.get('iDisplayStart')
            consulta = consulta[int(iDisplayStart):]
        if request.GET.get('iDisplayLength'):
            iDisplayLength = request.GET.get('iDisplayLength')
            consulta = consulta[:int(iDisplayLength)]
        serializer = self.serializer(consulta, many=True)
        data = {
            "aaData": serializer.data,
            "iTotalRecords": iTotalRecords,
            "sEcho": sEcho,
            "iTotalDisplayRecords": iTotalDisplayRecords
        }
        return HttpResponse(json.dumps(data), content_type="application/json")


class DatosInactivos(DatosActivos):
    activo = 0


class EmpleadoList(TemplateView):
    template_name = "empleado/empleado_list.html"


class EmpleadoRecycle(TemplateView):
    template_name = "empleado/empleado_recycle.html"


class EmpleadoFormView(FormView):
    model = Empleado
    second_model = Direccion
    third_model = DatosUsuarioEmpleado
    fourth_model = User
    fifth_model = PasswordCliente
    form_class = EmpleadoForm
    second_form_class = DireccionForm
    third_form_class = RegistrarEditarForm
    success_url = reverse_lazy('empleado:empleado_list')
    template_name = "empleado/empleado_form.html"

    def get(self, request, *args, **kwargs):
        # Aqui se coloco el get de forma en que si se desea editar un registro inactivo redireccione a la lista
        id_empleado = kwargs.get('pk')
        url = self.request.resolver_match.url_name
        form = self.get_form()
        form2 = self.second_form_class
        form3 = self.third_form_class
        num_empleado_usuario = self.request.user.username
        if id_empleado is not None:
            try:
                # Busca si existe el dato
                datos = self.third_model.objects.get(id=id_empleado)
                if datos.numero_empleado == num_empleado_usuario:
                    # Redirecciona si el propio administrador intenta editar sus datos en esta vista
                    raise SuspiciousOperation('400')
                elif datos.activo == 0:
                    # Redirecciona si es activo 0
                    raise Http404
                else:
                    # Renderiza si esta activo
                    return self.render_to_response(self.get_context_data(form=form, form2=form2, form3=form3))
            except self.third_model.DoesNotExist:
                raise Http404
        else:
            # Redirecciona si con el pk enviado no existe un dato
            # Si intenta usar editar sin pk
            if url == 'empleado_editar':
                raise SuspiciousOperation('400')
                # Si es crear
            elif url == 'empleado_nuevo':
                # Renderiza porque si pk es none se trata de un nuevo registro
                return self.render_to_response(self.get_context_data(form=form, form2=form2, form3=form3))

    def get_context_data(self, **kwargs):
        context = super(EmpleadoFormView, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk', 0)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class()
        if 'form3' not in context:
            context['form3'] = self.third_form_class()
        context['id'] = pk
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        pais = pais_default
        id_empleado = kwargs.get('pk', 0)
        tsp = transaction.savepoint()
        try:
            if id_empleado == 0:
                request.POST._mutable = True
                numero_empleado = convertirnumeroempleadoentero(int(request.POST[u'numero_empleado_entero']))
                request.POST[u'numero_empleado'] = numero_empleado
                request.POST._mutable = False
                grupo = self.request.POST.get('grupo', 0)
                form = self.form_class(request.POST)
                form2 = self.second_form_class(request.POST)
                form3 = self.third_form_class(request.POST)
                if form.is_valid() and form2.is_valid() and form3.is_valid():
                    # creando el empleado
                    empleado = form.save()
                    # creando la direccion
                    direccion = form2.save()
                    direccion.pais = pais
                    direccion.save()
                    # relacionando el empleado con la direccion
                    empleado.direccion = direccion
                    empleado.save()
                    # creando el datosusuarioempleado
                    datos = form3.save()
                    # creando el usuario
                    email = self.request.POST.get('email')
                    usuario = self.fourth_model(username=str(int(datos.numero_empleado)), email=email)
                    usuario.first_name = empleado.nombre
                    if empleado.apellido_materno == '':
                        usuario.last_name = empleado.apellido_paterno
                    else:
                        usuario.last_name = empleado.apellido_paterno + ' ' + empleado.apellido_materno
                    usuario.set_password(datos.numero_empleado)
                    passwordhasheada = hashlib.md5(datos.numero_empleado).hexdigest()
                    password = self.fifth_model(password=passwordhasheada)
                    password.save()
                    grupos = traergrupos(grupo)
                    usuario.save()
                    usuario.groups.clear()
                    for grupo in grupos:
                        usuario.groups.add(grupo)
                    usuario.save()
                    # relacionando el usuario y el empleado con los datos
                    datos.password = password
                    datos.usuario = usuario
                    datos.empleado = empleado
                    datos.save()
                    if tsp:
                        transaction.savepoint_commit(tsp)
                    return HttpResponseRedirect(self.get_success_url())
                else:
                    return self.render_to_response(
                        self.get_context_data(form=form, form2=form2, form3=form3))
            else:
                request.POST._mutable = True
                numero_empleado = convertirnumeroempleadoentero(int(request.POST[u'numero_empleado_entero']))
                request.POST[u'numero_empleado'] = numero_empleado
                request.POST._mutable = False
                grupo = self.request.POST.get('grupo', 0)
                datos = self.third_model.objects.get(id=id_empleado)
                cambiar_password = False
                if datos.usuario.check_password(datos.numero_empleado):
                    cambiar_password = True
                password = datos.password
                empleado = self.model.objects.get(id=datos.empleado.id)
                direccion = self.second_model.objects.get(id=datos.empleado.direccion.id)
                username = datos.usuario.username
                user = self.fourth_model.objects.get(username=username)
                form = self.form_class(request.POST, instance=empleado)
                form2 = self.second_form_class(request.POST, instance=direccion)
                form3 = self.third_form_class(request.POST, instance=datos)
                if form.is_valid() and form2.is_valid() and form3.is_valid():
                    # guadando cambios en empleado y direccion
                    empleado = form.save()
                    direccion = form2.save()
                    direccion.pais = pais
                    direccion.save()
                    empleado.direccion = direccion
                    empleado.save()
                    # guardando cambios en datos
                    datos = form3.save()
                    # guadando cambios en usuario
                    if cambiar_password:
                        # Si el número de empleado cambia tambien cambian las contraseñas
                        user.set_password(datos.numero_empleado)
                        passwordhasheada = hashlib.md5(datos.numero_empleado).hexdigest()
                        password.password = passwordhasheada
                        password.save()
                    user.username = str(int(datos.numero_empleado))
                    user.first_name = empleado.nombre
                    user.email = self.request.POST.get('email')
                    if empleado.apellido_materno == '':
                        user.last_name = empleado.apellido_paterno
                    else:
                        user.last_name = empleado.apellido_paterno + ' ' + empleado.apellido_materno
                    grupos = traergrupos(grupo)
                    user.groups.clear()
                    for grupo in grupos:
                        user.groups.add(grupo)
                    user.save()
                    # guadando cambios en datos y reestableciendo el numero de empleado en caso de haber sido cambiado
                    datos.password = password
                    datos.usuario = user
                    datos.empleado = empleado
                    datos.save()
                    if tsp:
                        transaction.savepoint_commit(tsp)
                    return HttpResponseRedirect(self.get_success_url())
                else:
                    return self.render_to_response(
                        self.get_context_data(form=form, form2=form2, form3=form3))
        except self.model.DoesNotExist:
            raise Exception('500')
        except self.second_model.DoesNotExist:
            raise Exception('500')
        except self.third_model.DoesNotExist:
            raise Exception('500')
        except self.fourth_model.DoesNotExist:
            raise Exception('500')
        except self.fifth_model.DoesNotExist:
            raise Exception('500')
        except Exception, e:
            raise e


class ImportarExcel(GuardarEmpleadoMixin, FormView):
    form_class = ImportarExcelForm
    success_url = reverse_lazy('empleado:empleado_list')
    template_name = 'empleado/importar_excel.html'

    def post(self, request, *args, **kwargs):
        response_data = {u'Exito': False, u'Error': None, u'Resultados': None, u'Errores': None}
        archivo = request.FILES[u'archivo']
        largo = len(archivo.name) - 5
        extension = archivo.name[largo:]
        if extension == u'.xlsx':
            workbook = load_workbook(archivo)
            hojas = workbook.worksheets
            for hoja in hojas:
                if hoja.max_column != 20:
                    response_data[u'Error'] = u'<strong>' \
                                              u'<p class="text-center text-danger">' \
                                              u'El archivo excel no tiene el formato requerido.' \
                                              u'</p>' \
                                              u'</strong>'
                elif hoja.max_column == 20:
                    response_data[u'Exito'] = True
                    filas = []
                    titulos = [u'Numero de Empleado (del 0 al 9999) *', u'Nombre *', u'Apellido Paterno *',
                               u'Apellido Materno', u'Correo Electronico *', u'Hora de Entrada (24 HRS HH:MM) *',
                               u'Hora de Salida (24 HRS HH:MM) *', u'Fecha de Nacimiento (AAAA-MM-DD)', u'CURP', u'RFC',
                               u'Pais', u'Estado', u'Municipio', u'Ciudad', u'Asentamiento', u'Calle',
                               u'Numero Exterior', u'Numero Interior', u'Codigo Postal', u'Datos Adicionales']
                    vacio = [None, None, None, None, None, None, None, None, None, None,
                             None, None, None, None, None, None, None, None, None, None]
                    for rows in hoja.iter_rows('A1:T' + str(hoja.max_row)):
                        row = []
                        for cell in rows:
                            row.append(cell.value)
                        filas.append(row)
                    contador = 0
                    errores = []
                    resultados = u''
                    for fila in filas:
                        if fila != titulos and fila != vacio:
                            guardado = u'fallido'
                            if fila[0] is not None:
                                str_numero_empleado = convertirnumeroempleadoentero(int(fila[0]))
                                try:
                                    DatosUsuarioEmpleado.objects.get(numero_empleado__iexact=str_numero_empleado)
                                    guardado = u'modificado'
                                except Exception:
                                    guardado = u'creado'
                            error = self.guardar(fila, request)
                            if error[u'NoGuardado']:
                                guardado = u'fallido'
                            resultados += unicode(self.resultado(self, contador, guardado, fila, error))
                            errores.append({u'id': unicode(contador)+u'_resultado', u'error': error[u'errores']})
                            contador += 1
                    response_data[u'Errores'] = errores
                    response_data[u'Resultados'] = resultados
        else:
            response_data[u'Error'] = u'<strong>' \
                                      u'<p class="text-center text-danger">' \
                                      u'El archivo cargado no es un excel (.xlsx).' \
                                      u'</p>' \
                                      u'</strong>'
        return HttpResponse(json.dumps(response_data), content_type="application/json")


class ExportarEmpleados(GenerarExcelMixin, View):

    def get(self, request, *args, **kwargs):
        nombre_archivo_excel = "Empleados"
        columnas_ancho = [
            (u"Numero de Empleado(del 0 al 9999) *", 40), (u"Nombre *", 20), (u"Apellido Paterno *", 20),
            (u"Apellido Materno", 20), (u"Correo Electronico *", 25), (u'Hora de Entrada (24 HRS HH:MM) *', 37),
            (u'Hora de Salida (24 HRS HH:MM) *', 37), (u'Fecha de Nacimiento (AAAA-MM-DD)', 40),
            (u'CURP', 24), (u'RFC', 14), (u'Pais', 10), (u'Estado', 10), (u'Municipio', 10), (u'Ciudad', 10),
            (u'Asentamiento', 20), (u'Calle', 20), (u'Numero Exterior', 17), (u'Numero Interior', 17),
            (u'Codigo Postal', 15), (u'Datos Adicionales', 20)
        ]
        fila_titulo = 'A1:T1'
        consulta = DatosUsuarioEmpleado.objects.all().values(
            'usuario__username', 'empleado__nombre', 'empleado__apellido_paterno', 'empleado__apellido_materno',
            'usuario__email', 'hora_entrada', 'hora_salida', 'empleado__fecha_nacimiento', 'empleado__curp',
            'empleado__rfc', 'empleado__direccion__pais', 'empleado__direccion__estado',
            'empleado__direccion__municipio', 'empleado__direccion__ciudad', 'empleado__direccion__asentamiento',
            'empleado__direccion__calle', 'empleado__direccion__numero_exterior',
            'empleado__direccion__numero_interior', 'empleado__direccion__codigo_postal',
            'empleado__direccion__datos_adicionales').order_by('numero_empleado').filter(activo=1)
        if consulta.count() > 0:
            return self.generarexcelresponse(nombre_archivo_excel, columnas_ancho, fila_titulo, consulta)
